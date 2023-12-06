# -----------------------------------------------------------------------------
# ------------------------------ LIBRERIA RIESGO ------------------------------
# -----------------------------------------------------------------------------
"""
-------------------------------------------------------------------------------
Este script contiene las funciones que permiten procesar los resultados de 
riesgo
---------------------------- Autor: Daniela Novoa -----------------------------
"""
#%% ====== Importar librerias =================================================
""" 
-------------------------------------------------------------------------------
*** Ubicar secciones para guardar las librerias
-------------------------------------------------------------------------------
"""
# -------- Librerias Interfaz>> Tkinter ---------------------------------------
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
# -------- Librerias Graficos Interfaz>> Tkinter ------------------------------
from PIL import Image, ImageTk
import matplotlib.ticker as ticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
# -------- Librerias Directorios ----------------------------------------------
import os
import glob
import zipfile
import io
# -------- Librerias procesamiento de datos -----------------------------------
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
# -------- Librerias para generar las tablas de resumen -----------------------
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
# -------- Librerias Graficos Interfaz>> Tkinter ------------------------------
import matplotlib.ticker as ticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
# -------- Generar mapas ------------------------------------------------------
import geopandas as gpd
import contextily as ctx
from matplotlib.colors import Normalize
import matplotlib.pyplot as plt
from matplotlib import cm
from mpl_toolkits.axes_grid1.anchored_artists import AnchoredSizeBar
import matplotlib.font_manager as fm
import matplotlib.offsetbox as offsetbox
from matplotlib.patches import Patch
from matplotlib.patches import Rectangle
import re
#%% ====== Funcion Exportar ===================================================
"""
-------------------------------------------------------------------------------
Funciones exportar archivos
-------------------------------------------------------------------------------
"""
def ExportarGraficos_Perdidas_Calibrar(canvas1, canvas2, canvase1, canvase2, dflist_sts, dflist_mnz, N_sim):
    # Pregunta al usuario dónde desea guardar los archivos PNG
    directorio_destino = filedialog.askdirectory(title="Select a directory to save the files")
    if directorio_destino:
        # Guardar el primer gráfico
        filename1 = os.path.join(directorio_destino, "CentroPoblado_Ses.svg")
        canvas1.figure.savefig(filename1, bbox_inches='tight', format='svg', dpi=300)
        # Guardar el segundo gráfico
        filename2 = os.path.join(directorio_destino, "Manzana_Ses.svg")
        canvas2.figure.savefig(filename2, bbox_inches='tight', format='svg', dpi=300)
        # Guardar el tercer grafico
        filename3 = os.path.join(directorio_destino, "CentroPoblado_events.svg")
        canvase1.figure.savefig(filename3, bbox_inches='tight', format='svg', dpi=300)
        # Guardar el cuarto gráfico
        filename4 = os.path.join(directorio_destino, "Manzana_events.svg")
        canvase2.figure.savefig(filename4, bbox_inches='tight', format='svg', dpi=300)
        # Guardar .csv de cada simulacion
        for ind, df_sts in enumerate(dflist_sts):
            nombre_carpeta = str(int(N_sim[ind])) + '_Events'
            ruta_carpeta = os.path.join(directorio_destino, nombre_carpeta)
            if not os.path.exists(ruta_carpeta):
                os.makedirs(ruta_carpeta)  # Crea el directorio si no existe
            df_sts.to_csv(os.path.join(ruta_carpeta, 'aggrisk-stats.csv'),index=False)
            dflist_mnz[ind].to_csv(os.path.join(ruta_carpeta, 'aggrisk-stats-cod_mnz.csv'),index=False)
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Export Graphics", "The files have been successfully saved in:\n\n" + directorio_destino)

def ExportarGraficos_Perdidas_Dispersion(canvas1, canvas2):
    # Pregunta al usuario dónde desea guardar los archivos PNG
    directorio_destino = filedialog.askdirectory(title="Select a directory to save the files")
    if directorio_destino:
        # Guardar el primer gráfico
        filename1 = os.path.join(directorio_destino, "CentroPoblado_Ses.svg")
        canvas1.figure.savefig(filename1, bbox_inches='tight', format='svg', dpi=300)
        # Guardar el segundo gráfico
        filename2 = os.path.join(directorio_destino, "CentroPoblado_Events.svg")
        canvas2.figure.savefig(filename2, bbox_inches='tight', format='svg', dpi=300)
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Export Graphics", "The files have been successfully saved in:\n\n" + directorio_destino)

def Exportar_Perdidas_RiskByevent(canvas1,dfexcel1,resumen_tax,canvas2):
    # Pregunta al usuario dónde desea guardar los archivos PNG
    directorio_destino = filedialog.askdirectory(title="Selecciona un directorio para guardar los archivos")
    if directorio_destino:
        # Guardar el primer gráfico
        filename1 = os.path.join(directorio_destino, "Curv_Excedencia_Promedio.svg")
        canvas1.figure.savefig(filename1, bbox_inches='tight', format='svg', dpi=300)
        filename2 = os.path.join(directorio_destino, "PAE_Municipio.xlsx")
        dfexcel1(filename2)  
        filename3 = os.path.join(directorio_destino, "PAE_Taxonomia.xlsx")
        resumen_tax(filename3) 
        filename4 = os.path.join(directorio_destino, "Diagrama_taxonomia.svg")
        canvas2.figure.savefig(filename4, bbox_inches='tight', format='svg', dpi=300)
        # Informar al usuario que los gráficos se han guardado con éxito
        tk.messagebox.showinfo("Exportar archivos", "Los archivos se han guardado con éxito en:\n\n" + directorio_destino)

#%% ====== Funcion generar tabla resumen EBR ==================================
"""
-------------------------------------------------------------------------------
Funcion generar tabla de resumen. perdidas>EBR
-------------------------------------------------------------------------------
"""
def gen_tabla(valorexp,PAE_mill,PE_mill):
    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active
    # Configuracion de la primera linea 
    ws.merge_cells('B2:C2')
    ws['B2'] = 'Valor expuesto'
    ws['D2'] = '[COP millones]'
    ws['E2'] = valorexp
    # Configuracion de la segunda linea 
    ws.merge_cells('B3:C4')
    ws['B3'] = 'Pérdida anual esperada'
    ws['D3'] = '[COP millones]'
    ws['E3'] = PAE_mill
    ws['D4'] = '[‰]'
    ws['E4'] = (PAE_mill/valorexp)*1000
    # Configuracion de la tercera linea
    ws.merge_cells('B5:E5')
    ws['B5'] = 'Pérdida máxima probable'
    # Configuracion de la tercera linea 
    ws.merge_cells('B6:B7')
    ws['B6'] = 'Periodo de retorno'
    ws.merge_cells('C6:C7')
    ws['C6'] = 'Probabilidad de excedencia en 50 años'
    ws.merge_cells('D6:E7')
    ws['D6'] = 'Pérdida esperada'
    # Configuracion de la cuarta linea
    ws['B8'] = '[años]'
    ws['C8'] = '[%]'
    ws['D8'] = '[COP millones]'
    ws['E8'] = '[%]'
    # Configuracion de la quinta linea
    ws['B9'] = 31
    ws['C9'] = (1-np.exp(-50/31))*100
    ws['D9'] = PE_mill[0]
    ws['E9'] = (PE_mill[0]/valorexp)*100
    # Configuracion de la sexta linea
    ws['B10'] = 225
    ws['C10'] = (1-np.exp(-50/225))*100
    ws['D10'] = PE_mill[1]
    ws['E10'] = (PE_mill[1]/valorexp)*100
    # Configuracion de la septima linea
    ws['B11'] = 475
    ws['C11'] = (1-np.exp(-50/475))*100
    ws['D11'] = PE_mill[2]
    ws['E11'] = (PE_mill[2]/valorexp)*100
    # Configuracion de la octava linea
    ws['B12'] = 975
    ws['C12'] = (1-np.exp(-50/975))*100
    ws['D12'] = PE_mill[3]
    ws['E12'] = (PE_mill[3]/valorexp)*100
    # Configuracion de la novena linea
    ws['B13'] = 1475
    ws['C13'] = (1-np.exp(-50/1475))*100
    ws['D13'] = PE_mill[4]
    ws['E13'] = (PE_mill[4]/valorexp)*100

    # Alinear el contenido de las celdas al centro
    for row in ws.iter_rows(min_row=2, max_row=13, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    # Configurar el tamaño de las celdas
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    # Agregar bordes a todas las celdas
    for row in ws.iter_rows(min_row=2, max_row=13, min_col=2, max_col=5):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
    # Guardar el archivo 
    Table_Resu = wb.save
    return Table_Resu
#%% ====== Funcion generar tabla resumen EBR Taxonomia ========================
"""
-------------------------------------------------------------------------------
Funcion generar tabla de resumen. perdidas>EBR
-------------------------------------------------------------------------------
"""
def gen_tabla_tax(df_expotax,taxo_description):
    wb = Workbook()
    ws = wb.active
    # Configuracion de la primera fila
    ws.merge_cells('B2:C2')
    ws['B2'] = 'Tipología constructiva'
    ws.merge_cells('D2:E2')
    ws['D2'] = 'Valor expuesto'
    ws.merge_cells('F2:H2')
    ws['F2'] = 'Pérdida anual esperada'
    # Configuración de la segunda fila
    ws['B3'] = 'Descripción'
    ws['C3'] = 'Taxonomía'
    ws['D3'] = '[COP millones]'
    ws['E3'] = '[%]'
    ws['F3'] = '[COP millones]'
    ws['G3'] = '[%]'
    ws['H3'] = '[‰]'
    # Configuracion taxonomia
    for index, txn in enumerate(df_expotax.taxonomy):
        ws['B'+str(index+4)] = taxo_description[index]
        ws['C'+str(index+4)] = txn
        ws['D'+str(index+4)] = np.around(df_expotax.valex[index],3)
        ws['E'+str(index+4)] = np.around((df_expotax.valex[index]/np.sum(df_expotax.valex))*100,3)
        ws['F'+str(index+4)] = np.around(df_expotax.loss[index],3)
        ws['G'+str(index+4)] = np.around((df_expotax.loss[index]/df_expotax.valex[index])*100,3)
        ws['H'+str(index+4)] = np.around((df_expotax.loss[index]/df_expotax.valex[index])*1000,3)

    # Alinear el contenido de las celdas al centro
    for row in ws.iter_rows(min_row=2, max_row=len(df_expotax.taxonomy)+3, min_col=2, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

    # Configurar el tamaño de las celdas
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 11

    # Agregar bordes a todas las celdas
    for row in ws.iter_rows(min_row=2, max_row=len(df_expotax.taxonomy)+3, min_col=2, max_col=8):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
    # Guardar el archivo 
    Table_Resu_tax = wb.save
    return Table_Resu_tax
#%% ====== Funcion tick modificados ===========================================
def format_tick(value,position):
    value_in_tik = int(value/1000)                                     
    return f'{value_in_tik}k'

def format_tick_EBR(value,position):
    value_in_tik = int(value)                                     
    return f'{value_in_tik}%'
#%% ====== Funcion generar graficos sin linea de tendencia / Calibrar =========
def canva_CLB_In(datos, xlabel, ylabel, title, canvas_master, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax1 = plt.subplots(figsize=(6, 3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    ax1.grid(True, lw=0.3, which='both')
    ax1.tick_params(labelsize=11, width=4)
    # ---- Procesar datos -----------------------------------------------------
    df = pd.DataFrame(datos)
    # ---- Configuración del gráfico ------------------------------------------    
    ax1.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax1.plot(df['Num_Sim'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax1.set_xlabel(xlabel, fontsize=12,fontweight='bold')
    ax1.set_ylabel(ylabel, fontsize=12, color='#262626',fontweight='bold')
    ax1.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    ax1.set_title(title)
    # ---- Añadir error -------------------------------------------------------
    ax2 = ax1.twinx()
    ax2.set_ylabel('Error [%]', color='#C34545',fontweight='bold', fontsize=12)
    ax2.plot(df['Num_Sim'][1:], df['error'][1:], '*', color='#C34545')
    numeromax = (int(np.max(df['error'].astype(float)) * 10) + 1) / 10
    intervalo = np.around((numeromax - 0.0) / (len(df['Num_Sim']) - 1),1)
    nuevos_ticks_y = np.around([0.0 + i * intervalo for i in range(len(df['Num_Sim']))],1)
    tiskpr = list(nuevos_ticks_y)
    tiskpr.append(2.0)
    ax2.tick_params(axis='y', labelcolor='#C34545')
    ax2.set_yticks(sorted(tiskpr))
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=600, height=430)
    return canvas
#%% ====== Funcion generar graficos con linea de tendencia / Calibrar =========
def canva_CLB_LT(datos, xlabel, ylabel, title, canvas_master, relex, reley):
    # datos: Diccinaro de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax = plt.subplots(figsize=(6, 3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    ax.grid(True, lw=0.3, which='both')
    ax.tick_params(labelsize=12, width=4)
    # ---- Procesar datos -----------------------------------------------------
    df = pd.DataFrame(datos)
    # ---- Configuración del gráfico ------------------------------------------    
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax.plot(df['Num_Sim'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax.set_xlabel(xlabel, fontsize=12,fontweight='bold')
    ax.set_ylabel(ylabel, fontsize=12, color='#262626',fontweight='bold')
    ax.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    ax.set_title(title)
    # ---- Añadir error -------------------------------------------------------
    ax2 = ax.twinx()
    ax2.set_ylabel('Error [%]', color='#C34545',fontweight='bold', fontsize=12)
    ax2.plot(df['Num_Sim'][1:], df['error'][1:], '*', color='#C34545')
    numeromax = (int(np.max(df['error'].astype(float)) * 10) + 1) / 10
    intervalo = (numeromax - 0.0) / (len(df['Num_Sim']) - 1)
    nuevos_ticks_y = np.around([0.0 + i * intervalo for i in range(len(df['Num_Sim']))],1)
    ax2.tick_params(axis='y', labelcolor='#C34545')
    ax2.set_yticks(nuevos_ticks_y)
    # ---- Linea regresion lineal ---------------------------------------------
    df['Num_Sim'] = df['Num_Sim'].astype(float)
    df['error'] = df['error'].astype(float)
    z = np.polyfit(df['Num_Sim'][1:], df['error'][1:], 1)
    p = np.poly1d(z)
    plt.plot(df['Num_Sim'], p(df['Num_Sim']), "--", color='#C34545')
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=600, height=430)
    return canvas
#%% ====== Funcion generar grafico / EBR ======================================
def canva_crv_EBR(datos, valex_entry, valper_entry, xlabel, ylabel, title, canvas_master, relex, reley):
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig = plt.figure(figsize=(6,3))
    fig, ax = plt.subplots()
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.12,right=0.98,top=0.98,bottom=0.11)
    ax.tick_params(labelsize=12, width=4)
    ax.grid(True, lw=0.15, which='both')
    # ---- Configuración del gráfico ------------------------------------------
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick_EBR))
    ax.plot((datos.loss/(valex_entry*1e6))*100,datos['perdidaTA'],color='#3B3838')
    if valper_entry is not None:
        anoslist = [475,975,2475,valper_entry]
        # color = ['#8064A2','#8CA448','#C0504D','#31859C','#BA7296','#B27E5E']
        color = ['#C0504D','#31859C','#BA7296','#B27E5E']
    else:
        anoslist = [475,975,2475]
        # anoslist = [31,225,475]
        # color = ['#8064A2','#8CA448','#C0504D','#31859C','#BA7296']
        color = ['#C0504D','#31859C','#BA7296']
    for inx,i in enumerate(anoslist):
        datos['diferencia'] = abs(datos['perdidaTA']-1/i)
        indice_point = datos['diferencia'].idxmin()
        perdidaval = (datos.loc[indice_point,'loss'])*100/(valex_entry*1e6)
        ax.plot([0,perdidaval],[1/i,1/i],'--',color=color[inx])
        ax.plot([perdidaval,perdidaval],[np.min(datos['perdidaTA']),1/i],'--',color=color[inx])
        ax.plot(perdidaval,1/i,'o',color=color[inx],markersize=4.0,label=f'{np.around(perdidaval,1)}% en {i} años')
    ax.set_xlabel(xlabel,fontsize=12)
    ax.set_ylabel(ylabel,fontsize=12)
    ax.set_title(title,fontsize=14)
    ax.set_xlim(np.min((datos.loss/(valex_entry*1e6))*100),np.max((datos.loss/(valex_entry*1e6))*100))
    ax.set_ylim(np.min(datos['perdidaTA']),np.max(datos['perdidaTA']))
    ax.legend()
    ax.set_yscale('log')
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH,expand=False)
    canvas.get_tk_widget().place(relx=relex,rely=reley,anchor=tk.CENTER, width=550, height=400)
    return canvas
#%% ====== Funcion generar diagrama taxonmia perdidas =========================
def canva_EBR_taxo(df_expotax,canvas_master,relex,reley):
    # Crear la figura y los ejes
    fig, ax1 = plt.subplots(figsize=(9,7))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.27)
    # Crear el primer gráfico de barras
    ax1.bar(np.array(range(len(df_expotax.taxonomy))) - 0.2, np.around(df_expotax.valex/1e6,3), width=0.4, label='Valor expuesto [COP Billones]', color='orange',alpha=0.5)
    ax1.set_ylabel('Valor expuesto [COP billones]', color='orange',fontsize=9)
    ax1.tick_params(axis='y', labelcolor='orange',labelsize=10)
    # Crear un segundo eje para el segundo gráfico de barras
    ax2 = ax1.twinx()
    ax2.bar(np.array(range(len(df_expotax.taxonomy))) + 0.2, np.around(df_expotax.loss,3)*0.001, width=0.4, label='Pérdida anual esperada [COP miles de millón]', color='blue',alpha=0.5)
    ax2.set_ylabel('Pérdida anual esperada [COP miles de millón]', color='blue',fontsize=9)
    ax2.tick_params(axis='y', labelcolor='blue',labelsize=10)
    # Añadir la línea de puntos con valores sobre cada punto
    line_data = np.around((df_expotax.loss/df_expotax.valex)*1000,3)
    for i in range(int(len(df_expotax.taxonomy)/2)):
        ax2.plot(np.array(range(len(df_expotax.taxonomy)))[i], line_data[i], 'ko',markersize=3.0)
        ax2.text(np.array(range(len(df_expotax.taxonomy)))[i]+0.54, line_data[i]-0.0015, f'{line_data[i]}‰', color='k', ha='center', fontsize=7)
    for i in range(int(len(df_expotax.taxonomy)/2),len(df_expotax.taxonomy)):
        ax2.plot(np.array(range(len(df_expotax.taxonomy)))[i], line_data[i], 'ko',markersize=3.0)
        ax2.text(np.array(range(len(df_expotax.taxonomy)))[i]+0.47, line_data[i]-0.002, f'{line_data[i]}‰', color='k', ha='center', fontsize=7)
    line, = ax2.plot(np.NaN, np.NaN, 'ko-', markersize=3.0, label='Pérdida anual esperada [‰]')
    ax2.plot(np.array(range(len(df_expotax.taxonomy))),line_data,'k-',linewidth=1, alpha = 0.6)
    x_labels = df_expotax.taxonomy
    # Cambiar los ticks del eje x a letras
    ax1.set_xticks(np.array(range(len(df_expotax.taxonomy))))
    ax1.set_xticklabels(x_labels,rotation=90,fontsize=9)
    # Añadir leyendas y título
    ax1.set_xlabel('Taxonomias',fontsize=9)
    fig.legend(loc='lower center', bbox_to_anchor=(0.5, -0.003), ncol=2, fontsize=8)
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH,expand=False)
    canvas.get_tk_widget().place(relx=relex,rely=reley,anchor=tk.CENTER, width=650, height=500)
    return canvas
#%% ====== Funcion generar represetacion espacial COP =========================
def canva_mapa_COP(map_data,seccion_shp,area_shpe,COD_mun,CP_Name,ruta_shp,canvas_master,relex,reley):
    # Generar mapa
    fig, (ax, ax_legend) = plt.subplots(1, 2, figsize=(9.5,6), gridspec_kw={'width_ratios': [1, 0.3]})
    # plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    # map_data.plot(column='loss', ax=ax, edgecolor='grey', alpha=0.7, cmap='Reds', linewidth=0.4)
    map_data.plot(column='loss', ax=ax, edgecolor='grey', alpha=0.7, cmap='Reds', linewidth=0.4)
    seccion_shp.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=0.5)
    area_shpe.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=0.5)
    # Brujula señalando el norte
    x, y, arrow_length = 0.95, 0.95, 0.1
    ax.annotate('N', xy=(x, y), xytext=(x, y-arrow_length),
                arrowprops=dict(facecolor='black', width=3, headwidth=9),
                ha='center', va='center', fontsize=11, xycoords=ax.transAxes)

    # Graficar:
    norm = Normalize(vmin=map_data['loss'].min(), vmax=map_data['loss'].max())
    cax = fig.add_axes([0.752, 0.285, 0.13, 0.015]) # [left, bottom, width, height]
    cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap='Reds'), cax=cax, orientation='horizontal')
    valores = np.linspace(map_data['loss'].min(), map_data['loss'].max(), 7)
    cbar.set_ticks((np.round(valores[1:-1] / 0.1) * 0.1)) #** modificar para la plataforma
    cbar.ax.tick_params(axis='x', labelsize=5.5)
    ax_legend.text(0.5, 0.25, 'PAE [COP Millones]', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)

    ax_legend.set_axis_off()

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"acofilogo.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0215)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.87), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"serviciogc.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0515)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.87), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)


    ax_legend.add_patch(Rectangle((0.736, 0.65), 0.159, 0.08, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.5, 0.76, 'Modelo Nacional', transform=ax_legend.transAxes,ha='center', va='center', fontsize=9)
    ax_legend.text(0.5, 0.73, 'de Riesgo Sísmico', transform=ax_legend.transAxes,ha='center', va='center', fontsize=9)

    ax_legend.add_patch(Rectangle((0.736, 0.56), 0.159, 0.08, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.5, 0.65, COD_mun, transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)
    ax_legend.text(0.5, 0.61, CP_Name, transform=ax_legend.transAxes,ha='center', va='center', fontsize=11)

    ax_legend.add_patch(Rectangle((0.736, 0.47), 0.159, 0.08, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.5, 0.52, 'Pérdida anual esperada', transform=ax_legend.transAxes,ha='center', va='center', fontsize=8)
    ax_legend.text(0.5, 0.49, 'en millones de pesos COP', transform=ax_legend.transAxes,ha='center', va='center', fontsize=8)

    ax_legend.add_patch(Rectangle((0.736, 0.25), 0.159, 0.21, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.235, 0.415, 'Leyenda', transform=ax_legend.transAxes,ha='center', va='center', fontsize=9)
    ax_legend.add_patch(Rectangle((0.745, 0.4), 0.015, 0.015, edgecolor='black', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.43, 0.373, 'Área censal', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)
    ax_legend.add_patch(Rectangle((0.745, 0.37), 0.015, 0.015, edgecolor='black', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.48, 0.333, 'Sección urbana', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)
    ax_legend.add_patch(Rectangle((0.745, 0.34), 0.015, 0.015, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.40, 0.293, 'Manzana', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"logosabana.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.17)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.11), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"logomedellin.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.12)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.11), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)

    # Añadir mapa base

    ctx.add_basemap(ax, crs=map_data.crs.to_string(), source=ctx.providers.CartoDB.Positron, zoom=13)

    datadf = pd.read_csv(os.path.join(ruta_shp,"Exposicion.csv")) 

    # Aplicar la función corregida a cada fila de la columna 'geometry' y crear una lista de todas las coordenadas
    all_coords_safe = []
    datadf['geometry'].apply(lambda x: all_coords_safe.extend(extract_coordinates_safe(x)))

    # Calcular los valores máximos y mínimos de longitud y latitud
    max_lat = max(all_coords_safe, key=lambda x: x[1])[1]
    min_lat = min(all_coords_safe, key=lambda x: x[1])[1]
    max_lon = max(all_coords_safe, key=lambda x: x[0])[0]
    min_lon = min(all_coords_safe, key=lambda x: x[0])[0]

    ax.set_xlim([np.ceil((min_lon-0.03) * 50)/50,np.floor((max_lon+0.03) * 50) / 50])
    ax.set_ylim([np.floor((min_lat-0.003) * 50) / 50, np.ceil((max_lat+0.005) * 50)/50])

    max_lat_new = np.ceil((min_lon-0.03) * 50)/50
    min_lat_new = np.floor((max_lon+0.03) * 50) / 50
    max_lon_new = np.floor((min_lat-0.003) * 50) / 50
    min_lon_new = np.ceil((max_lat+0.005) * 50)/50

    ticks_x = np.arange(min_lat_new, max_lat_new  - 0.02, -0.02)
    ax.set_xticks(ticks_x)
    ticks_y = np.arange(min_lon_new, max_lon_new  - 0.0, -0.01)
    ax.set_yticks(ticks_y)

    suma = 0
    len_marginx = 0.2 #*** modificar para la plataforma
    for index in range(len(ticks_x)-1):
        if suma%2 == 0:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 0.99), len_marginx, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.01, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 0.99), len_marginx, 0.01, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1

    suma = 0
    len_marginy = 0.125 #*** modificar para la plataforma
    for index in range(len(ticks_y)-1):
        if suma%2 == 0:
            rect = Rectangle((0.0, len_marginy*suma), 0.01, len_marginy, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((0.99, len_marginy*suma), 0.01, len_marginy, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((0.0, len_marginy*suma), 0.01, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((0.99, len_marginy*suma), 0.01, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1

    # Generar grilla
    ax.grid(True, which='both', color='grey', linewidth=1.0, linestyle='-', alpha=0.4)

    # rect = Rectangle((0.006, 0.013), 0.25, 0.12, fill=True, color='white', transform=ax.transAxes, clip_on=False)
    # ax.add_patch(rect)
    rect = Rectangle((0.07, 0.06), 0.125/(0.125/len_marginx), 0.01, fill=True, edgecolor='black', facecolor='none', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.125/(0.125/len_marginx)+0.07*1.3, 0.09, '2km', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.19, 0.06), (0.125/(0.125/len_marginx))/5, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.23, 0.09, '1.5', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.105, 0.06), (0.125/(0.125/len_marginx))/5, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.19, 0.09, '1', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.087, 0.06), 0.05*0.125/(0.125/len_marginx), 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.145, 0.09, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.07, 0.06), 0.05*0.125/(0.125/len_marginx), 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.105, 0.09, '0', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    ax.text(0.07, 0.09, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=6)

    ax.tick_params(axis='both', which='major', labelsize=8)
    plt.subplots_adjust(wspace=0.005)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH,expand=False)
    canvas.get_tk_widget().place(relx=relex,rely=reley,anchor=tk.CENTER, width=960, height=500)
    
    return canvas
#%% ====== Funcion generar represetacion espacial %. =========================
def canva_mapa_prc(map_data,seccion_shp,area_shpe,COD_mun,CP_Name,ruta_shp,canvas_master,relex,reley):
    # Generar mapa
    fig, (ax, ax_legend) = plt.subplots(1, 2, figsize=(9.5,6), gridspec_kw={'width_ratios': [1, 0.3]})
    # plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    # map_data.plot(column='loss', ax=ax, edgecolor='grey', alpha=0.7, cmap='Reds', linewidth=0.4)
    map_data.plot(column='loss2', ax=ax, edgecolor='grey', alpha=0.7, cmap='Reds', linewidth=0.4)
    seccion_shp.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=0.5)
    area_shpe.plot(ax=ax, edgecolor='black', facecolor="none", alpha=1.0, linewidth=0.5)
    # Brujula señalando el norte
    x, y, arrow_length = 0.95, 0.95, 0.1
    ax.annotate('N', xy=(x, y), xytext=(x, y-arrow_length),
                arrowprops=dict(facecolor='black', width=3, headwidth=9),
                ha='center', va='center', fontsize=11, xycoords=ax.transAxes)

    # Graficar:
    norm = Normalize(vmin=map_data['loss2'].min(), vmax=map_data['loss2'].max())
    cax = fig.add_axes([0.752, 0.285, 0.13, 0.015]) # [left, bottom, width, height]
    cbar = fig.colorbar(cm.ScalarMappable(norm=norm, cmap='Reds'), cax=cax, orientation='horizontal')
    valores = np.linspace(map_data['loss2'].min(), map_data['loss2'].max(), 7)
    cbar.set_ticks((np.round(valores[1:-1] / 0.01) * 0.01)) #** modificar para la plataforma
    cbar.ax.tick_params(axis='x', labelsize=4.5)
    ax_legend.text(0.5, 0.25, 'PAE [‰]', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)

    ax_legend.set_axis_off()

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"acofilogo.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0215)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.87), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"serviciogc.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.0515)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.87), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)


    ax_legend.add_patch(Rectangle((0.736, 0.65), 0.159, 0.08, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.5, 0.76, 'Modelo Nacional', transform=ax_legend.transAxes,ha='center', va='center', fontsize=9)
    ax_legend.text(0.5, 0.73, 'de Riesgo Sísmico', transform=ax_legend.transAxes,ha='center', va='center', fontsize=9)

    ax_legend.add_patch(Rectangle((0.736, 0.56), 0.159, 0.08, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.5, 0.65, COD_mun, transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)
    ax_legend.text(0.5, 0.61, CP_Name, transform=ax_legend.transAxes,ha='center', va='center', fontsize=11)

    ax_legend.add_patch(Rectangle((0.736, 0.47), 0.159, 0.08, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.5, 0.52, 'Pérdida anual esperada [‰]', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7.0)
    ax_legend.text(0.5, 0.49, '(pérdida/valor expuesto)', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7.0)

    ax_legend.add_patch(Rectangle((0.736, 0.25), 0.159, 0.21, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.235, 0.415, 'Leyenda', transform=ax_legend.transAxes,ha='center', va='center', fontsize=9)
    ax_legend.add_patch(Rectangle((0.745, 0.4), 0.015, 0.015, edgecolor='black', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.43, 0.373, 'Área censal', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)
    ax_legend.add_patch(Rectangle((0.745, 0.37), 0.015, 0.015, edgecolor='black', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.48, 0.333, 'Sección urbana', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)
    ax_legend.add_patch(Rectangle((0.745, 0.34), 0.015, 0.015, edgecolor='grey', facecolor='none', transform=fig.transFigure))
    ax_legend.text(0.40, 0.293, 'Manzana', transform=ax_legend.transAxes,ha='center', va='center', fontsize=7)

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"logosabana.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.17)
    ab = offsetbox.AnnotationBbox(imagebox, (0.25, 0.11), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)

    image_path = os.path.join(os.path.join(os.getcwd(),"icon"),"logomedellin.png")
    img = plt.imread(image_path)
    imagebox = offsetbox.OffsetImage(img, zoom=0.12)
    ab = offsetbox.AnnotationBbox(imagebox, (0.75, 0.11), frameon=False, xycoords='axes fraction', boxcoords="axes fraction")
    ax_legend.add_artist(ab)

    # Añadir mapa base

    ctx.add_basemap(ax, crs=map_data.crs.to_string(), source=ctx.providers.CartoDB.Positron, zoom=13)

    datadf = pd.read_csv(os.path.join(ruta_shp,"Exposicion.csv")) 

    # Aplicar la función corregida a cada fila de la columna 'geometry' y crear una lista de todas las coordenadas
    all_coords_safe = []
    datadf['geometry'].apply(lambda x: all_coords_safe.extend(extract_coordinates_safe(x)))

    # Calcular los valores máximos y mínimos de longitud y latitud
    max_lat = max(all_coords_safe, key=lambda x: x[1])[1]
    min_lat = min(all_coords_safe, key=lambda x: x[1])[1]
    max_lon = max(all_coords_safe, key=lambda x: x[0])[0]
    min_lon = min(all_coords_safe, key=lambda x: x[0])[0]

    ax.set_xlim([np.ceil((min_lon-0.03) * 50)/50,np.floor((max_lon+0.03) * 50) / 50])
    ax.set_ylim([np.floor((min_lat-0.003) * 50) / 50, np.ceil((max_lat+0.005) * 50)/50])

    max_lat_new = np.ceil((min_lon-0.03) * 50)/50
    min_lat_new = np.floor((max_lon+0.03) * 50) / 50
    max_lon_new = np.floor((min_lat-0.003) * 50) / 50
    min_lon_new = np.ceil((max_lat+0.005) * 50)/50

    ticks_x = np.arange(min_lat_new, max_lat_new  - 0.02, -0.02)
    ax.set_xticks(ticks_x)
    ticks_y = np.arange(min_lon_new, max_lon_new  - 0.0, -0.01)
    ax.set_yticks(ticks_y)

    suma = 0
    len_marginx = 0.2 #*** modificar para la plataforma
    for index in range(len(ticks_x)-1):
        if suma%2 == 0:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 0.99), len_marginx, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((len_marginx*suma, 0.0), len_marginx, 0.01, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((len_marginx*suma, 0.99), len_marginx, 0.01, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1

    suma = 0
    len_marginy = 0.125 #*** modificar para la plataforma
    for index in range(len(ticks_y)-1):
        if suma%2 == 0:
            rect = Rectangle((0.0, len_marginy*suma), 0.01, len_marginy, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((0.99, len_marginy*suma), 0.01, len_marginy, fill=True, color='black', transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1
        else:
            rect = Rectangle((0.0, len_marginy*suma), 0.01, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            rect = Rectangle((0.99, len_marginy*suma), 0.01, len_marginy, fill=True, edgecolor='black', facecolor='white', linewidth=0.8, transform=ax.transAxes, clip_on=False)
            ax.add_patch(rect)
            suma = suma + 1

    # Generar grilla
    ax.grid(True, which='both', color='grey', linewidth=1.0, linestyle='-', alpha=0.4)

    # rect = Rectangle((0.006, 0.013), 0.25, 0.12, fill=True, color='white', transform=ax.transAxes, clip_on=False)
    # ax.add_patch(rect)
    rect = Rectangle((0.07, 0.06), 0.125/(0.125/len_marginx), 0.01, fill=True, edgecolor='black', facecolor='none', linewidth=0.8, transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.125/(0.125/len_marginx)+0.07*1.3, 0.09, '2km', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.19, 0.06), (0.125/(0.125/len_marginx))/5, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.23, 0.09, '1.5', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.105, 0.06), (0.125/(0.125/len_marginx))/5, 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.19, 0.09, '1', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.087, 0.06), 0.05*0.125/(0.125/len_marginx), 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.145, 0.09, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    rect = Rectangle((0.07, 0.06), 0.05*0.125/(0.125/len_marginx), 0.01, fill=True, color='black', transform=ax.transAxes, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.105, 0.09, '0', transform=ax.transAxes,ha='center', va='center', fontsize=6)
    ax.text(0.07, 0.09, '0.5', transform=ax.transAxes,ha='center', va='center', fontsize=6)

    ax.tick_params(axis='both', which='major', labelsize=8)
    plt.subplots_adjust(wspace=0.005)
    
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH,expand=False)
    canvas.get_tk_widget().place(relx=relex,rely=reley,anchor=tk.CENTER, width=960, height=500)
    
    return canvas
#%% ====== Funcion limites mapa ===============================================
def extract_coordinates_safe(polygon_str):
    """
    Extracts the coordinates from a POLYGON string and returns them as a list of tuples (latitude, longitude).
    """
    if not isinstance(polygon_str, str):
        # Si la entrada no es una cadena, la convierte en una
        polygon_str = str(polygon_str)

    # Encuentra todas las coincidencias de patrones de coordenadas en la cadena del polígono
    coords = re.findall(r'(-?\d+\.\d+) (-?\d+\.\d+)', polygon_str)
    # Convierte las coordenadas a flotantes y las guarda en tuplas (longitud, latitud)
    return [(float(lon), float(lat)) for lon, lat in coords]
#%% ====== Funcion generar graficos por eventos generados =====================
def canva_events(datos, xlabel, ylabel, title, canvas_master, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax = plt.subplots(figsize=(6, 4.3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    ax.grid(True, lw=0.3, which='both')
    ax.tick_params(labelsize=12, width=4)
    # ---- Procesar datos -----------------------------------------------------
    df = pd.DataFrame(datos)
    # ---- Configuración del gráfico ------------------------------------------
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax.plot(df['Num_Sim'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax.set_xlabel(xlabel, fontsize=12,fontweight='bold')
    ax.set_ylabel(ylabel, fontsize=12, color='#262626',fontweight='bold')
    ax.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    ax.set_title(title)
    # ---- Añadir error -------------------------------------------------------
    ax2 = ax.twinx()
    ax2.set_ylabel('Error [%]', color='#C34545',fontweight='bold', fontsize=12)
    ax2.plot(df['Num_Sim'][1:], df['error'][1:], '*', color='#C34545')
    numeromax = (int(np.max(df['error'].astype(float)) * 10) + 1) / 10
    intervalo = (numeromax - 0.0) / (len(df['Num_Sim']) - 1)
    nuevos_ticks_y = np.around([0.0 + i * intervalo for i in range(len(df['Num_Sim']))],1)
    tiskpr = list(nuevos_ticks_y)
    tiskpr.append(1.0)
    ax2.tick_params(axis='y', labelcolor='#C34545')
    ax2.set_yticks(sorted(tiskpr))
    return fig
#%% ====== Funcion generar graficos / DSP =====================================
def canva_DSP(datos, xlabel, ylabel, title, canvas_master, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax1 = plt.subplots(figsize=(6, 3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    ax1.grid(True, lw=0.3, which='both')
    ax1.tick_params(labelsize=11, width=4)
    # ---- Procesar datos -----------------------------------------------------
    df = pd.DataFrame(datos)
    df.sort_values(by='Num_Sim',ascending=True,inplace=True)
    # ---- Configuración del gráfico ------------------------------------------    
    ax1.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax1.plot(df['Num_Sim'], df['loss'], 'o--', color='#262626', markersize=4.5)
    ax1.set_xlabel(xlabel, fontsize=12,fontweight='bold')
    ax1.set_ylabel(ylabel, fontsize=12, color='#262626',fontweight='bold')
    ax1.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    ax1.set_title(title)
    # ---- Crear lienzo para mostrar la gráfica en la interfaz ----------------
    canvas = FigureCanvasTkAgg(fig, master=canvas_master)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=False)
    canvas.get_tk_widget().place(relx=relex, rely=reley, anchor=tk.CENTER, width=600, height=430)
    return canvas

#%% ====== Funcion generar graficos / DSP events ==============================
def canva_DSP_Ev(datos, xlabel, ylabel, title, canvas_master, relex, reley):
    # datos: Diccionario de datos
    # canvas_master: Frame en donde estara ubicado el grafico
    # ---- Parametros de la grafica -------------------------------------------
    plt.rcParams['savefig.bbox'] = 'tight'
    plt.style.use('default')
    plt.rcParams['font.family'] = 'Calibri'
    fig, ax1 = plt.subplots(figsize=(6, 4.3))
    fig.set_facecolor('white')
    plt.subplots_adjust(left=0.15, right=0.85, top=0.90, bottom=0.20)
    ax1.grid(True, lw=0.3, which='both')
    ax1.tick_params(labelsize=11, width=4)
    # ---- Procesar datos -----------------------------------------------------
    df = pd.DataFrame(datos)
    # ---- Configuración del gráfico ------------------------------------------    
    ax1.xaxis.set_major_formatter(ticker.FuncFormatter(format_tick))
    ax1.plot(df['Num_Sim'], df['loss'], 'o', color='#262626', markersize=4.5)
    ax1.set_xlabel(xlabel, fontsize=12,fontweight='bold')
    ax1.set_ylabel(ylabel, fontsize=12, color='#262626',fontweight='bold')
    ax1.tick_params(axis='y', labelcolor='#262626', labelsize=10)
    ax1.set_title(title)
    return fig